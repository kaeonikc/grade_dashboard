#!/usr/bin/env python3
import argparse
import sys
import shutil
import yaml
import pandas as pd
import datetime
import re
from pathlib import Path
from rich.prompt import Confirm

# Ensure src module can be imported, even if run from a softlink
script_path = Path(__file__).resolve()
sys.path.insert(0, str(script_path.parent))

from src.data_loader import load_config, load_course_data, sync_attendance_xlsx_to_csv
from src.calculators import calculate_final_grades

def create_merged_config(metadata: dict, output_path: Path):
    config_data = {
        'course_id': metadata.get('course_id'),
        'course_name': metadata.get('course_name'),
        'term': metadata.get('term'),
        'sec_num': metadata.get('sec_num'),
        'teacher': metadata.get('teacher'),
        'credits': metadata.get('credits'),
        'university': metadata.get('university'),
        'campus': metadata.get('campus'),
        'program': metadata.get('program'),
        'class_schedule': metadata.get('class_schedule'),
        'exam_schedule': metadata.get('exam_schedule'),
        'term_start_date': '',
        'weights': {
            'attendance': 0.10,
            'homework': 0.20,
            'midterm': 0.30,
            'final': 0.40
        },
        'data_mapping': {
            'homework': ['hw1', 'hw2'],
            'midterm': ['midterm_score'],
            'final': ['final_exam']
        },
        'grade_boundaries': {
            'A': 80,
            'B+': 75,
            'B': 70,
            'C+': 65,
            'C': 60,
            'D+': 55,
            'D': 50
        }
    }
    with open(output_path, 'w', encoding='utf-8') as f:
        yaml.dump(config_data, f, allow_unicode=True, default_flow_style=False, sort_keys=False)

THAI_WEEKDAY_MAP = {
    'จ.': 0,    # Monday
    'อ.': 1,    # Tuesday
    'พ.': 2,    # Wednesday
    'พฤ.': 3,   # Thursday
    'ศ.': 4,    # Friday
    'ส.': 5,    # Saturday
    'อา.': 6    # Sunday
}

MONTH_ABBRS = {
    1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun',
    7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'
}

def parse_midterm_date(exam_schedule_str: str) -> datetime.date | None:
    if not exam_schedule_str or not isinstance(exam_schedule_str, str):
        return None
    parts = exam_schedule_str.split(',')
    for part in parts:
        part = part.strip()
        if not re.search(r'\bM\b', part):
            continue
        
        # Try matching "d mmm yyyy" format: e.g., 7 Aug 2026
        match_new = re.search(r'\b(\d{1,2})\s+([A-Za-z]{3})\s+(\d{4})\b', part)
        if match_new:
            day = int(match_new.group(1))
            month_str = match_new.group(2)
            year = int(match_new.group(3))
            
            MONTH_MAP = {
                'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
                'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
            }
            m_num = MONTH_MAP.get(month_str.lower())
            if m_num:
                try:
                    return datetime.date(year, m_num, day)
                except ValueError:
                    pass
                    
        # Try matching "DD/MM/YY" format
        match_old = re.search(r'\b(\d{2})/(\d{2})/(\d{2,4})\b', part)
        if match_old:
            try:
                d = int(match_old.group(1))
                m = int(match_old.group(2))
                y_str = match_old.group(3)
                year = 2000 + int(y_str) if len(y_str) == 2 else int(y_str)
                return datetime.date(year, m, d)
            except ValueError:
                pass
    return None

def calculate_class_dates(term_start_date_str: str, weekday_str: str, exam_schedule_str: str, num_classes: int = 15) -> list[str]:
    """
    Calculates class dates based on term start date, weekday of the class,
    and skips the week of the midterm exam.
    Returns list of formatted date strings "d mmm yyyy" (e.g. "7 Aug 2026").
    """
    if not term_start_date_str:
        return [f"a{i}" for i in range(1, num_classes + 1)]
        
    try:
        start_date = datetime.datetime.strptime(term_start_date_str.strip(), "%Y-%m-%d").date()
    except Exception:
        return [f"a{i}" for i in range(1, num_classes + 1)]
        
    weekday_int = THAI_WEEKDAY_MAP.get(weekday_str.strip(), 0)
    
    days_ahead = weekday_int - start_date.weekday()
    if days_ahead < 0:
        days_ahead += 7
    first_class_date = start_date + datetime.timedelta(days=days_ahead)
    
    midterm_date = parse_midterm_date(exam_schedule_str)
    midterm_iso_week = None
    midterm_year = None
    if midterm_date:
        midterm_year, midterm_iso_week, _ = midterm_date.isocalendar()
        
    class_dates = []
    curr_date = first_class_date
    while len(class_dates) < num_classes:
        curr_year, curr_iso_week, _ = curr_date.isocalendar()
        
        if midterm_iso_week is not None and curr_iso_week == midterm_iso_week and curr_year == midterm_year:
            curr_date += datetime.timedelta(days=7)
            continue
            
        formatted = f"{curr_date.day} {MONTH_ABBRS[curr_date.month]} {curr_date.year}"
        class_dates.append(formatted)
        curr_date += datetime.timedelta(days=7)
        
    return class_dates

def save_attendance_excel(excel_path: Path, student_registry: pd.DataFrame, columns: list[str]):
    """
    Creates or updates the attendance spreadsheet using openpyxl.
    Maintains existing scores if they exist.
    Applies dropdown data validations, conditional colors, rotated headers, and compact layout.
    """
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter

    existing_df = None
    if excel_path.exists():
        try:
            existing_df = pd.read_excel(excel_path)
            existing_df.columns = existing_df.columns.astype(str).str.strip()
        except Exception as e:
            print(f"⚠️ Warning: Could not read existing attendance file: {e}")

    # Build header formulas. We keep using the static value for the maximum score (total number of class sessions).
    if len(columns) > 0:
        header_formula = f"total ({len(columns)}pts)"
    else:
        header_formula = "total (0pts)"
    
    headers = ['Student ID', 'Name'] + columns + [header_formula]
    
    df_new = pd.DataFrame(columns=headers)
    df_new['Student ID'] = student_registry['Student ID'].astype(str).str.strip()
    df_new['Name'] = student_registry['Name'].astype(str).str.strip()
    
    # Merge existing data if available
    if existing_df is not None and 'Student ID' in existing_df.columns:
        existing_df['Student ID'] = existing_df['Student ID'].astype(str).str.strip()
        old_cols = [c for c in existing_df.columns if c not in ['Student ID', 'Name'] and not c.lower().startswith('total') and not c.startswith('=') and 'unnamed' not in c.lower()]
        new_cols = columns
        
        student_rows = {sid: idx for idx, sid in enumerate(df_new['Student ID'])}
        
        for idx, sid in enumerate(existing_df['Student ID']):
            if sid in student_rows:
                target_idx = student_rows[sid]
                for pos, old_col in enumerate(old_cols):
                    if pos < len(new_cols):
                        new_col = new_cols[pos]
                        if old_col in new_cols:
                            val = existing_df.loc[idx, old_col]
                            if pd.notna(val):
                                df_new.loc[target_idx, old_col] = val
                        else:
                            val = existing_df.loc[idx, old_col]
                            if pd.notna(val):
                                df_new.loc[target_idx, new_col] = val

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.views.sheetView[0].showGridLines = True
    
    ws.append(headers)
    for r_idx, row in df_new.iterrows():
        ws.append(row.tolist())
        
    num_students = len(df_new)
    # Date columns start at column 3 (C) and end at column len(columns)+2 (e.g. Q)
    # Date columns start at column 3 (C) and end at column len(columns)+2 (e.g. Q)
    # The total column is at column len(columns)+3
    if len(columns) > 0:
        start_col_letter = get_column_letter(3)
        end_col_letter = get_column_letter(len(columns) + 2)
        total_col_idx = len(columns) + 3
        for r in range(2, num_students + 2):
            formula = f'=COUNTIF({start_col_letter}{r}:{end_col_letter}{r}, "P") + COUNTIF({start_col_letter}{r}:{end_col_letter}{r}, "L")*0.8 + COUNTIF({start_col_letter}{r}:{end_col_letter}{r}, "EA")'
            ws.cell(row=r, column=total_col_idx, value=formula)
        
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    ws.row_dimensions[1].height = 75
    header_font = Font(name='Segoe UI', size=11, bold=True)
    normal_font = Font(name='Segoe UI', size=11)
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.border = thin_border
        
        if col_idx > 2 and col_idx < len(headers):
            # Rotated for date/placeholder columns only
            cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='center', wrap_text=True)
        elif col_idx == len(headers):
            # Total header is not rotated
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
    for r in range(2, num_students + 2):
        ws.row_dimensions[r].height = 20
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = normal_font
            cell.border = thin_border
            if c == len(headers):
                cell.font = Font(name='Segoe UI', size=11, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif c > 2:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    for c in range(3, len(headers)):
        col_letter = get_column_letter(c)
        ws.column_dimensions[col_letter].width = 6
    total_col_letter = get_column_letter(len(headers))
    ws.column_dimensions[total_col_letter].width = 14
        
    dv = DataValidation(type="list", formula1='"P,A,L,EA"', allow_blank=True)
    dv.error ='Your entry is not in the list (P, A, L, EA)'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = 'Please select from: P, A, L, EA'
    dv.promptTitle = 'Select attendance'
    
    ws.add_data_validation(dv)
    
    if len(columns) > 0 and num_students > 0:
        start_col_letter = get_column_letter(3)
        end_col_letter = get_column_letter(len(columns) + 2)
        validation_range = f"{start_col_letter}2:{end_col_letter}{num_students + 1}"
        dv.add(validation_range)
        
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="006100", name='Segoe UI', size=11)
    
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006", name='Segoe UI', size=11)
    
    orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    orange_font = Font(color="9C6500", name='Segoe UI', size=11)
    
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    yellow_font = Font(color="808000", name='Segoe UI', size=11)
    
    if len(columns) > 0 and num_students > 0:
        cf_range = f"{start_col_letter}2:{end_col_letter}{num_students + 1}"
        ws.conditional_formatting.add(cf_range, CellIsRule(operator='equal', formula=['"P"'], fill=green_fill, font=green_font))
        ws.conditional_formatting.add(cf_range, CellIsRule(operator='equal', formula=['"A"'], fill=red_fill, font=red_font))
        ws.conditional_formatting.add(cf_range, CellIsRule(operator='equal', formula=['"L"'], fill=orange_fill, font=orange_font))
        ws.conditional_formatting.add(cf_range, CellIsRule(operator='equal', formula=['"EA"'], fill=yellow_fill, font=yellow_font))

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)
    print(f"✅ Generated and formatted attendance sheet at: {excel_path.name}")

def parse_pts(col_name: str) -> tuple[str, float | None]:
    """
    Parses a column name to see if it already contains points notation.
    Returns (clean_name, pts_value) or (col_name, None).
    """
    import re
    match = re.search(r'\(\s*([\d.]+)\s*(?:pts|points)?\s*\)', col_name, re.IGNORECASE)
    if match:
        try:
            val = float(match.group(1))
            clean_name = re.sub(r'\s*\(\s*[\d.]+\s*(?:pts|points)?\s*\)', '', col_name, flags=re.IGNORECASE).strip()
            return clean_name, val
        except ValueError:
            pass
    return col_name, None

def parse_config_col(col) -> tuple[str, float | None]:
    """
    Parses a column item from config file (can be a string or a dictionary).
    Returns (clean_name, pts_value) or (col_name_str, None).
    """
    if isinstance(col, dict):
        if len(col) == 1:
            key = list(col.keys())[0]
            val = col[key]
            try:
                return str(key).strip(), float(val)
            except (ValueError, TypeError):
                return str(key).strip(), None
        return str(col), None
    else:
        col_str = str(col).strip()
        return parse_pts(col_str)

def parse_category_mapping(category_mapping) -> tuple[float | None, list]:
    """
    Parses a category mapping from the configuration.
    It can be a list of columns, or a dictionary with 'total_pts' and 'columns'.
    Returns (total_pts, columns_list).
    """
    if isinstance(category_mapping, dict):
        total_pts = category_mapping.get('total_pts')
        if total_pts is not None:
            try:
                total_pts = float(total_pts)
            except (ValueError, TypeError):
                total_pts = None
        columns_list = category_mapping.get('columns', [])
        return total_pts, columns_list
    else:
        return None, category_mapping

def create_empty_category_csv(category: str, columns: list, student_registry: pd.DataFrame, csv_path: Path, default_max: float = 100.0):
    headers = ['Student ID', 'Name']
    total_pts = 0.0
    for col in columns:
        clean_name, pts = parse_config_col(col)
        pts_val = pts if pts is not None else default_max
        total_pts += pts_val
        
        pts_str = str(int(pts_val)) if pts_val.is_integer() else str(pts_val)
        headers.append(f"{clean_name} ({pts_str}pts)")
        
    # Append the total column header
    total_pts_str = str(int(total_pts)) if total_pts.is_integer() else str(total_pts)
    headers.append(f"total ({total_pts_str}pts)")
    
    student_rows = student_registry.copy()
    for col in headers:
        if col not in student_rows.columns:
            student_rows[col] = ""
    student_rows = student_rows[headers]
    
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    student_rows.to_csv(csv_path, index=False)

def init_course(course_name: str, term_name: str, course_id: str, input_file: str = None):
    if input_file:
        input_path = Path(input_file)
        if not input_path.exists():
            print(f"❌ Input file not found: {input_path}")
            sys.exit(1)
        try:
            # Check file type and read
            if input_path.suffix.lower() == '.csv':
                df_raw = pd.read_csv(input_path, header=None)
            else:
                df_raw = pd.read_excel(input_path, header=None)
        except Exception as e:
            print(f"❌ Error reading raw file: {e}")
            sys.exit(1)
        
        from course_info_prep.convert_student_info import extract_metadata
        metadata = extract_metadata(df_raw)
        
        term_name = term_name if term_name else metadata.get('term')
        course_id = course_id if course_id else metadata.get('course_id')
        course_name = course_name if course_name else metadata.get('course_name')
        sec_num = metadata.get('sec_num', 'Unknown')
        
        folder_name = f"{term_name}_{course_id}_SEC_{sec_num}_grading"
    else:
        folder_name = f"{term_name}_{course_id}_grading"
        sec_num = None

    base_dir = Path(folder_name)
    if base_dir.exists():
        if input_file:
            print(f"❌ Directory '{base_dir}' already exists. Use 'grader update -i {input_file}' to update it.")
        else:
            print(f"❌ Directory '{base_dir}' already exists. Aborting to avoid overwriting existing files.")
        sys.exit(1)

    course_info_dir = base_dir / "course_info"
    data_dir = base_dir / "data"
    reports_dir = base_dir / "reports"
    
    course_info_dir.mkdir(parents=True)
    data_dir.mkdir(parents=True)
    reports_dir.mkdir(parents=True)

    if input_file:
        config_filename = f"{term_name}_{course_id}_SEC_{sec_num}_config.yaml"
        config_path = course_info_dir / config_filename
        csv_filename = f"{term_name}_{course_id}_SEC_{sec_num}_student_info.csv"
        csv_path = course_info_dir / csv_filename
        
        temp_yaml_path = course_info_dir / "temp_metadata.yaml"
        from course_info_prep.convert_student_info import convert_excel_to_csv_and_yaml
        metadata, clean_df = convert_excel_to_csv_and_yaml(input_path, csv_path=csv_path, yaml_path=temp_yaml_path)
        
        create_merged_config(metadata, config_path)
        if temp_yaml_path.exists():
            temp_yaml_path.unlink()
        
        # Generate initial attendance spreadsheet with placeholders
        att_filename = f"{term_name}_{course_id}_attendance.xlsx"
        att_path = data_dir / att_filename
        class_cols = calculate_class_dates('', '', '')
        save_attendance_excel(att_path, clean_df, class_cols)
        sync_attendance_xlsx_to_csv(att_path, att_path.with_suffix(".csv"))
        
        print(f"✅ Successfully created course structure at: {base_dir}")
        print(f"✅ Generated student list at: {csv_path}")
        print(f"✅ Generated config at: {config_path}")
        print(f"✅ Generated initial attendance sheet at: {att_path}")
    else:
        config_filename = f"{term_name}_{course_id}_config.yaml"
        config_path = course_info_dir / config_filename
        
        with open(config_path, "w") as f:
            f.write(f"""course_id: "{course_id}"
course_name: "{course_name}"
term: "{term_name}"
term_start_date: ""

weights:
  attendance: 0.10
  homework: 0.20
  midterm: 0.30
  final: 0.40

data_mapping:
  homework: ["hw1", "hw2"]
  midterm: ["midterm_score"]
  final: ["final_exam"]

grade_boundaries:
  A: 80
  B+: 75
  B: 70
  C+: 65
  C: 60
  D+: 55
  D: 50
""")
        print(f"✅ Successfully created course structure at: {base_dir}")
        print(f"Modify the '{config_path}' file to set up your rules!")

def mkdb_course(config_file: str = None, input_alias: str = None):
    cfg_path = None
    if config_file:
        cfg_path = Path(config_file)
    elif input_alias:
        cfg_path = Path(input_alias)
    else:
        # Auto-detect if inside course folder
        info_dir = Path("course_info")
        if info_dir.is_dir():
            config_files = [f for f in info_dir.iterdir() if f.is_file() and f.name.endswith("_config.yaml") and not f.name.endswith(".bak")]
            if config_files:
                cfg_path = config_files[0]
                print(f"🔍 Auto-detected configuration file: {cfg_path}")
                
    if not cfg_path or not cfg_path.exists():
        print("❌ Error: Configuration file not specified or not found.")
        print("Please provide the path to the config file or run the command from inside the course directory.")
        sys.exit(1)
        
    course_dir = cfg_path.resolve().parent.parent
    info_dir = course_dir / "course_info"
    data_dir = course_dir / "data"
    
    # Load config yaml
    try:
        with open(cfg_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f) or {}
    except Exception as e:
        print(f"❌ Error loading config YAML: {e}")
        sys.exit(1)
        
    data_mapping = config.get("data_mapping", {})
    if not data_mapping:
        print(f"⚠️ Warning: 'data_mapping' not found or empty in {cfg_path.name}.")
        return
        
    # Locate student info CSV
    student_info_files = [f for f in info_dir.iterdir() if f.is_file() and f.name.endswith("_student_info.csv") and not f.name.endswith(".bak")]
    if not student_info_files:
        print(f"❌ Error: Student registry (*_student_info.csv) not found under {info_dir}")
        sys.exit(1)
        
    student_info_path = student_info_files[0]
    try:
        student_df = pd.read_csv(student_info_path)
        student_df.columns = student_df.columns.astype(str).str.strip()
    except Exception as e:
        print(f"❌ Error reading student registry: {e}")
        sys.exit(1)
        
    if 'Student ID' not in student_df.columns or 'Name' not in student_df.columns:
        print(f"❌ Error: Registry must contain 'Student ID' and 'Name' columns.")
        sys.exit(1)
        
    # Clean registry
    student_registry = student_df[['Student ID', 'Name']].copy()
    student_registry['Student ID'] = student_registry['Student ID'].astype(str).str.strip()
    student_registry['Name'] = student_registry['Name'].astype(str).str.strip()
    
    data_dir.mkdir(parents=True, exist_ok=True)
    
    for category, mapping_item in data_mapping.items():
        total_pts, columns = parse_category_mapping(mapping_item)
        if not isinstance(columns, list):
            continue
        csv_path = data_dir / f"{category}.csv"
        print(f"🛠️ Creating empty score database: {csv_path.name}...")
        
        # Calculate target_total based on category weight or manual total_pts
        weights = config.get("weights", {})
        weight = weights.get(category, None)
        weight_points = weight * 100.0 if weight is not None else 100.0
        
        target_total = total_pts if total_pts is not None else weight_points
        
        specified_total = 0.0
        unspecified_cols = 0
        for col in columns:
            _, pts = parse_config_col(col)
            if pts is not None:
                specified_total += pts
            else:
                unspecified_cols += 1
                
        if unspecified_cols > 0:
            remaining = target_total - specified_total
            if remaining <= 0:
                default_max = 10.0 # Fallback default
            else:
                default_max = round(remaining / unspecified_cols, 2)
        else:
            default_max = 100.0
            # Warning check
            import math
            if not math.isclose(specified_total, target_total, abs_tol=1e-2):
                print(f"⚠️ WARNING: [{category}] Manual total score is determined as {target_total}pts, "
                      f"but the sum of specified columns is {specified_total}pts.")
        
        try:
            create_empty_category_csv(category, columns, student_registry, csv_path, default_max)
            print(f"✅ Created {csv_path.name} successfully with default max score {default_max}.")
        except Exception as e:
            print(f"❌ Error writing {csv_path.name}: {e}")
            
    print("🚀 All databases created successfully!")

def update_course(input_file: str = None, config_file: str = None):
    # 1. Raw excel file update mode (if input_file is supplied)
    if input_file:
        input_path = Path(input_file)
        if not input_path.exists():
            print(f"❌ Input file not found: {input_path}")
            sys.exit(1)
        try:
            if input_path.suffix.lower() == '.csv':
                df_raw = pd.read_csv(input_path, header=None)
            else:
                df_raw = pd.read_excel(input_path, header=None)
        except Exception as e:
            print(f"❌ Error reading raw file: {e}")
            sys.exit(1)
        
        from course_info_prep.convert_student_info import extract_metadata, convert_excel_to_csv_and_yaml
        metadata = extract_metadata(df_raw)
        term_name = metadata.get('term')
        course_id = metadata.get('course_id')
        sec_num = metadata.get('sec_num')
        
        if not term_name or not course_id or not sec_num:
            print("❌ Error: Could not parse required metadata (term, course_id, sec_num) from input file.")
            sys.exit(1)
            
        # Automatic detection of target folder
        target_dir = None
        if Path("course_info").is_dir():
            target_dir = Path(".")
            print("🔍 Detected running from inside a course grading directory.")
        else:
            folder_name = f"{term_name}_{course_id}_SEC_{sec_num}_grading"
            target_dir = Path(folder_name)
            # Check inside courses/ as well
            if not target_dir.exists() and Path("courses").is_dir():
                target_dir = Path("courses") / folder_name
                
        if not target_dir.exists():
            print(f"❌ Target grading directory '{target_dir}' does not exist.")
            print(f"Please run 'grader init -i {input_file}' first to initialize the course.")
            sys.exit(1)
            
        info_dir = target_dir / "course_info"
        if not info_dir.is_dir():
            print(f"❌ '{info_dir}' is not a directory.")
            sys.exit(1)
            
        # Determine config and CSV paths
        config_files = [f for f in info_dir.iterdir() if f.is_file() and f.name.endswith("_config.yaml") and not f.name.endswith(".bak")]
        if config_files:
            config_path = config_files[0]
        else:
            config_path = info_dir / f"{term_name}_{course_id}_SEC_{sec_num}_config.yaml"
            
        csv_path = info_dir / f"{term_name}_{course_id}_SEC_{sec_num}_student_info.csv"
        
        # Back up existing files
        if config_path.exists():
            shutil.copy2(config_path, config_path.with_suffix(".yaml.bak"))
            print(f"📦 Backed up configuration to: {config_path.with_suffix('.yaml.bak').name}")
        if csv_path.exists():
            shutil.copy2(csv_path, csv_path.with_suffix(".csv.bak"))
            print(f"📦 Backed up student info to: {csv_path.with_suffix('.csv.bak').name}")
            
        # Merge student info CSV
        if csv_path.exists():
            try:
                old_df = pd.read_csv(csv_path)
                old_df.columns = old_df.columns.astype(str).str.strip()
            except Exception as e:
                print(f"⚠️ Warning: Could not read old student CSV, overwriting instead: {e}")
                old_df = pd.DataFrame(columns=['Student ID', 'Name', 'Class Group'])
        else:
            old_df = pd.DataFrame(columns=['Student ID', 'Name', 'Class Group'])
            
        temp_csv = info_dir / "temp_new_students.csv"
        temp_yaml = info_dir / "temp_new_metadata.yaml"
        try:
            _, new_df = convert_excel_to_csv_and_yaml(input_path, csv_path=temp_csv, yaml_path=temp_yaml)
        finally:
            if temp_csv.exists(): temp_csv.unlink()
            if temp_yaml.exists(): temp_yaml.unlink()
            
        new_df['Student ID'] = new_df['Student ID'].astype(str).str.strip()
        old_df['Student ID'] = old_df['Student ID'].astype(str).str.strip()
        
        merged_df = pd.merge(old_df, new_df, on='Student ID', how='outer', suffixes=('_old', '_new'))
        merged_df['Name'] = merged_df['Name_new'].fillna(merged_df['Name_old']).fillna('')
        merged_df['Class Group'] = merged_df['Class Group_new'].fillna(merged_df['Class Group_old']).fillna('')
        merged_df = merged_df[['Student ID', 'Name', 'Class Group']]
        merged_df = merged_df.sort_values(by='Student ID').reset_index(drop=True)
        
        merged_df.to_csv(csv_path, index=False)
        print(f"✅ Merged student list. Total students now: {len(merged_df)} (Added {len(merged_df) - len(old_df)} new students)")
        
        # Merge config metadata
        if config_path.exists():
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    existing_config = yaml.safe_load(f) or {}
            except Exception as e:
                print(f"⚠️ Error reading existing config, starting fresh: {e}")
                existing_config = {}
                
            metadata_keys = [
                'course_id', 'course_name', 'term', 'sec_num', 'teacher', 
                'credits', 'university', 'campus', 'program', 'class_schedule', 'exam_schedule'
            ]
            for key in metadata_keys:
                if key in metadata:
                    existing_config[key] = metadata[key]
                    
            if 'term_start_date' not in existing_config:
                existing_config['term_start_date'] = ''
                
            if 'weights' not in existing_config:
                existing_config['weights'] = {'attendance': 0.10, 'homework': 0.20, 'midterm': 0.30, 'final': 0.40}
            else:
                if 'attendance' not in existing_config['weights']:
                    existing_config['weights']['attendance'] = 0.10
                    if 'final' in existing_config['weights']:
                        existing_config['weights']['final'] = round(existing_config['weights']['final'] - 0.10, 2)
                        
            if 'data_mapping' not in existing_config:
                existing_config['data_mapping'] = {'homework': ['hw1', 'hw2'], 'midterm': ['midterm_score'], 'final': ['final_exam']}
            if 'grade_boundaries' not in existing_config:
                existing_config['grade_boundaries'] = {'A': 80, 'B+': 75, 'B': 70, 'C+': 65, 'C': 60, 'D+': 55, 'D': 50}
                
            with open(config_path, 'w', encoding='utf-8') as f:
                yaml.dump(existing_config, f, allow_unicode=True, default_flow_style=False, sort_keys=False)
            print(f"✅ Merged metadata into configuration at: {config_path.name}")
            
            # Update attendance spreadsheet
            try:
                term_val = existing_config.get('term', term_name)
                cid_val = existing_config.get('course_id', course_id)
                att_filename = f"{term_val}_{cid_val}_attendance.xlsx"
                att_path = target_dir / "data" / att_filename
                
                sch = existing_config.get('class_schedule', {})
                weekday = sch.get('day', '') if isinstance(sch, dict) else ''
                exam_sch = existing_config.get('exam_schedule', '')
                term_start = existing_config.get('term_start_date', '')
                
                class_cols = calculate_class_dates(term_start, weekday, exam_sch)
                save_attendance_excel(att_path, merged_df, class_cols)
                sync_attendance_xlsx_to_csv(att_path, att_path.with_suffix(".csv"))
            except Exception as e:
                print(f"⚠️ Warning: Could not update attendance spreadsheet: {e}")
        else:
            create_merged_config(metadata, config_path)
            print(f"✅ Created configuration at: {config_path.name}")
            
            # Create fresh attendance spreadsheet with placeholders
            try:
                att_filename = f"{term_name}_{course_id}_attendance.xlsx"
                att_path = target_dir / "data" / att_filename
                class_cols = calculate_class_dates('', '', '')
                save_attendance_excel(att_path, merged_df, class_cols)
                sync_attendance_xlsx_to_csv(att_path, att_path.with_suffix(".csv"))
            except Exception as e:
                print(f"⚠️ Warning: Could not create attendance spreadsheet: {e}")
            
        print(f"🚀 Course update completed successfully!")
        return

    # 2. Configuration file update mode (if config_file is supplied or auto-detected)
    cfg_path = None
    if config_file:
        cfg_path = Path(config_file)
    else:
        # Auto-detect if inside course folder
        info_dir = Path("course_info")
        if info_dir.is_dir():
            config_files = [f for f in info_dir.iterdir() if f.is_file() and f.name.endswith("_config.yaml") and not f.name.endswith(".bak")]
            if config_files:
                cfg_path = config_files[0]
                print(f"🔍 Auto-detected configuration file: {cfg_path}")
                
    if not cfg_path or not cfg_path.exists():
        print("❌ Error: Configuration file not specified or not found.")
        print("Please provide the path to the config file or run update from inside a course grading directory.")
        sys.exit(1)
        
    course_dir = cfg_path.resolve().parent.parent
    info_dir = course_dir / "course_info"
    data_dir = course_dir / "data"
    
    # Load config yaml
    try:
        with open(cfg_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f) or {}
    except Exception as e:
        print(f"❌ Error loading config YAML: {e}")
        sys.exit(1)
        
    data_mapping = config.get("data_mapping", {})
    if not data_mapping:
        print(f"⚠️ Warning: 'data_mapping' not found or empty in {cfg_path.name}.")
        return
        
    # Locate student info CSV
    student_info_files = [f for f in info_dir.iterdir() if f.is_file() and f.name.endswith("_student_info.csv") and not f.name.endswith(".bak")]
    if not student_info_files:
        print(f"❌ Error: Student registry (*_student_info.csv) not found under {info_dir}")
        sys.exit(1)
        
    student_info_path = student_info_files[0]
    try:
        registry_df = pd.read_csv(student_info_path)
        registry_df.columns = registry_df.columns.astype(str).str.strip()
    except Exception as e:
        print(f"❌ Error reading student registry: {e}")
        sys.exit(1)
        
    if 'Student ID' not in registry_df.columns or 'Name' not in registry_df.columns:
        print(f"❌ Error: Registry must contain 'Student ID' and 'Name' columns.")
        sys.exit(1)
        
    # Clean registry
    student_registry = registry_df[['Student ID', 'Name']].copy()
    student_registry['Student ID'] = student_registry['Student ID'].astype(str).str.strip()
    student_registry['Name'] = student_registry['Name'].astype(str).str.strip()
    registry_ids = set(student_registry['Student ID'].tolist())
    
    data_dir.mkdir(parents=True, exist_ok=True)
    from rich.prompt import Confirm
    
    for category, mapping_item in data_mapping.items():
        total_pts, columns = parse_category_mapping(mapping_item)
        if not isinstance(columns, list):
            continue
        csv_path = data_dir / f"{category}.csv"
        
        # Calculate target_total based on category weight or manual total_pts
        weights = config.get("weights", {})
        weight = weights.get(category, None)
        weight_points = weight * 100.0 if weight is not None else 100.0
        
        target_total = total_pts if total_pts is not None else weight_points
        
        specified_total = 0.0
        unspecified_cols = 0
        for col in columns:
            _, pts = parse_config_col(col)
            if pts is not None:
                specified_total += pts
            else:
                unspecified_cols += 1
                
        if unspecified_cols > 0:
            remaining = target_total - specified_total
            if remaining <= 0:
                default_max = 10.0 # Fallback default
            else:
                default_max = round(remaining / unspecified_cols, 2)
        else:
            default_max = 100.0
            # Warning check
            import math
            if not math.isclose(specified_total, target_total, abs_tol=1e-2):
                print(f"⚠️ WARNING: [{category}] Manual total score is determined as {target_total}pts, "
                      f"but the sum of specified columns is {specified_total}pts.")
        
        # If database CSV doesn't exist, create it
        if not csv_path.exists():
            print(f"✨ Score database {csv_path.name} does not exist. Creating it...")
            create_empty_category_csv(category, columns, student_registry, csv_path, default_max)
            continue
            
        print(f"🔄 Aligning database score file: {csv_path.name}...")
        try:
            df = pd.read_csv(csv_path)
            df.columns = df.columns.astype(str).str.strip()
        except Exception as e:
            print(f"❌ Error reading {csv_path.name}: {e}. Skipping.")
            continue
            
        if 'Student ID' not in df.columns:
            print(f"❌ Error: {csv_path.name} is missing 'Student ID' column. Skipping.")
            continue
            
        # Separate Max/sentinel row from students (discard Max row if it exists)
        df['Student ID'] = df['Student ID'].astype(str).str.strip()
        max_mask = df['Student ID'].str.lower().isin(['max', 'max score', 'full score', 'full'])
        student_rows = df[~max_mask].copy()
        
        # Parse current headers to extract base names and original formatting
        csv_columns_info = {} # base_name -> (original_header, max_score)
        for col in df.columns:
            if col in ['Student ID', 'Name']:
                continue
            clean_name, pts = parse_pts(col)
            if clean_name.lower().strip() == 'total':
                continue
            if pts is not None:
                csv_columns_info[clean_name] = (col, pts)
            else:
                csv_columns_info[clean_name] = (col, 100.0)
                
        # Determine columns to add and remove
        desired_base_cols = []
        desired_base_mapping = {} # clean_name -> original_config_col
        for col in columns:
            clean_name, _ = parse_config_col(col)
            desired_base_cols.append(clean_name)
            desired_base_mapping[clean_name] = col
        
        # Columns to remove
        cols_to_remove = [b for b in csv_columns_info.keys() if b not in desired_base_cols]
        for col in cols_to_remove:
            orig_header = csv_columns_info[col][0]
            if orig_header in student_rows.columns:
                non_empty = student_rows[orig_header].dropna()
                non_empty = non_empty[non_empty.astype(str).str.strip() != '']
                non_empty = non_empty[non_empty != 0]
                
                if not non_empty.empty:
                    warn_msg = f"⚠️ WARNING: Column '{orig_header}' in '{csv_path.name}' contains student scores. Removing it will discard these scores. Do you want to proceed?"
                    if not Confirm.ask(warn_msg, default=False):
                        print("🚫 Aborted. Database update cancelled.")
                        sys.exit(0)
                student_rows.drop(columns=[orig_header], inplace=True)
                
        # Determine header mapping for desired columns
        final_headers = ['Student ID', 'Name']
        total_pts = 0.0
        for col in desired_base_cols:
            # Calculate the target new header name based on current config
            config_col_item = desired_base_mapping[col]
            clean_name, pts = parse_config_col(config_col_item)
            pts_val = pts if pts is not None else default_max
            total_pts += pts_val
            
            if pts_val is not None:
                pts_str = str(int(pts_val)) if pts_val.is_integer() else str(pts_val)
                new_header = f"{clean_name} ({pts_str}pts)"
            else:
                pts_str = str(int(default_max)) if default_max.is_integer() else str(default_max)
                new_header = f"{col} ({pts_str}pts)"

            if col in csv_columns_info:
                # Column exists, check if points changed and update header if needed
                orig_header = csv_columns_info[col][0]
                if orig_header != new_header:
                    student_rows.rename(columns={orig_header: new_header}, inplace=True)
                    print(f"  📝 Updated max points header for '{col}' from '{orig_header}' to '{new_header}'")
                final_headers.append(new_header)
            else:
                # Add new column header with specified or default max points
                student_rows[new_header] = ""
                final_headers.append(new_header)
                
        # Handle the total column at the end
        total_pts_str = str(int(total_pts)) if total_pts.is_integer() else str(total_pts)
        total_header = f"total ({total_pts_str}pts)"
        
        # Locate if there is an existing total column in student_rows to rename/align
        existing_total_col = None
        for col in student_rows.columns:
            if str(col).lower().strip().startswith('total'):
                existing_total_col = col
                break
                
        if existing_total_col:
            if existing_total_col != total_header:
                student_rows.rename(columns={existing_total_col: total_header}, inplace=True)
                print(f"  📝 Updated max points header for total column from '{existing_total_col}' to '{total_header}'")
        else:
            student_rows[total_header] = ""
            
        final_headers.append(total_header)
                
        # Students to remove
        csv_student_ids = student_rows['Student ID'].tolist()
        students_to_remove = [sid for sid in csv_student_ids if sid not in registry_ids]
        
        for sid in students_to_remove:
            stud_row = student_rows[student_rows['Student ID'] == sid]
            has_scores = False
            score_cols = [c for c in student_rows.columns if c not in ['Student ID', 'Name']]
            for c in score_cols:
                val = stud_row.iloc[0][c]
                if pd.notna(val) and str(val).strip() != '' and str(val).strip() != '0' and val != 0:
                    has_scores = True
                    break
                    
            if has_scores:
                student_name = stud_row.iloc[0].get('Name', 'Unknown')
                warn_msg = f"⚠️ WARNING: Student '{sid}' ({student_name}) has scores in '{csv_path.name}' but is no longer in the registry. Removing them will discard these scores. Do you want to proceed?"
                if not Confirm.ask(warn_msg, default=False):
                    print("🚫 Aborted. Database update cancelled.")
                    sys.exit(0)
            student_rows = student_rows[student_rows['Student ID'] != sid]
            
        # Students to add
        students_to_add = [sid for sid in registry_ids if sid not in csv_student_ids]
        add_rows = []
        for sid in students_to_add:
            reg_row = student_registry[student_registry['Student ID'] == sid]
            name = reg_row.iloc[0]['Name'] if not reg_row.empty else ""
            new_row = {col: "" for col in final_headers}
            new_row['Student ID'] = sid
            new_row['Name'] = name
            add_rows.append(new_row)
            
        if add_rows:
            student_rows = pd.concat([student_rows, pd.DataFrame(add_rows)], ignore_index=True)
            
        # Order columns and sort students
        student_rows = student_rows[final_headers]
        student_rows = student_rows.sort_values(by='Student ID').reset_index(drop=True)
        
        # Backup before writing
        shutil.copy2(csv_path, csv_path.with_suffix(".csv.bak"))
        print(f"  📦 Backed up {csv_path.name} to {csv_path.with_suffix('.csv.bak').name}")
        
        # Save
        student_rows.to_csv(csv_path, index=False)
        print(f"  ✅ Successfully updated {csv_path.name}")
        
    # Align/update attendance sheet if attendance is in weights
    if 'weights' in config and 'attendance' in config['weights']:
        try:
            term_val = config.get('term', '')
            cid_val = config.get('course_id', '')
            att_filename = f"{term_val}_{cid_val}_attendance.xlsx"
            att_path = data_dir / att_filename
            
            sch = config.get('class_schedule', {})
            weekday = sch.get('day', '') if isinstance(sch, dict) else ''
            exam_sch = config.get('exam_schedule', '')
            term_start = config.get('term_start_date', '')
            
            class_cols = calculate_class_dates(term_start, weekday, exam_sch)
            save_attendance_excel(att_path, student_registry, class_cols)
            sync_attendance_xlsx_to_csv(att_path, att_path.with_suffix(".csv"))
        except Exception as e:
            print(f"⚠️ Warning: Could not update/align attendance spreadsheet: {e}")
            
    print("🚀 All databases updated successfully!")

def undo_course():
    # Detect target folder
    target_dir = None
    if Path("course_info").is_dir():
        target_dir = Path(".")
    else:
        # Scan current dir and courses/ for folders with course_info/*.bak files
        potential_dirs = []
        search_paths = [Path(".")]
        if Path("courses").is_dir():
            search_paths.append(Path("courses"))
            
        for sp in search_paths:
            for d in sp.iterdir():
                if d.is_dir() and (d / "course_info").is_dir():
                    info_dir = d / "course_info"
                    bak_files = [f for f in info_dir.iterdir() if f.is_file() and f.name.endswith(".bak")]
                    if bak_files:
                        # Get latest modification time of any bak file
                        mtime = max(f.stat().st_mtime for f in bak_files)
                        potential_dirs.append((mtime, d))
                        
        if potential_dirs:
            # Pick the one with the most recently modified bak files
            potential_dirs.sort(key=lambda x: x[0], reverse=True)
            target_dir = potential_dirs[0][1]
            
    if not target_dir:
        print("❌ No backup files found to undo.")
        sys.exit(1)
        
    info_dir = target_dir / "course_info"
    bak_files = [f for f in info_dir.iterdir() if f.is_file() and f.name.endswith(".bak")]
    
    if not bak_files:
        print(f"❌ No backup files found in {target_dir / 'course_info'}.")
        sys.exit(1)
        
    print(f"🔄 Found backup files in course: {target_dir.name}")
    for bak in bak_files:
        original_name = bak.name[:-4] # strip .bak
        original_path = info_dir / original_name
        # Restore
        shutil.copy2(bak, original_path)
        bak.unlink()
        print(f"  ⏪ Restored {original_name} from backup")
    print("✅ Undo completed successfully!")

def run_dashboard():
    from src.dashboard import run
    run()

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="🎓 Grading System Dashboard & Manager")
    subparsers = parser.add_subparsers(dest="command", help="Available commands")

    # Init command
    parser_init = subparsers.add_parser("init", help="Initialize a new course grading directory")
    parser_init.add_argument("-i", "--input", help="Path to raw excel/csv file to initialize from (e.g. repclasslist.xls)")
    parser_init.add_argument("--course_name", help="Course display/short name (e.g. 'General Relativity')")
    parser_init.add_argument("--term", help="The term for the course (e.g. '2026_S2')")
    parser_init.add_argument("--course_id", help="The course ID (e.g. 'PHYS1102')")

    # mkdb command
    parser_mkdb = subparsers.add_parser("mkdb", help="Create empty database files for all categories from config")
    parser_mkdb.add_argument("config_file", nargs="?", help="Path to course config YAML file")
    parser_mkdb.add_argument("-i", "--input", help="Alias for path to course config YAML file")

    # Update command
    parser_update = subparsers.add_parser("update", help="Update student list from raw file or update database files from config")
    parser_update.add_argument("-i", "--input", help="Path to raw excel/csv student list file (e.g. repclasslist.xls)")
    parser_update.add_argument("config_file", nargs="?", help="Path to course config YAML file")

    # Undo command
    parser_undo = subparsers.add_parser("undo", help="Undo the last update by restoring backup files")

    # Dashboard command
    parser_dashboard = subparsers.add_parser("dashboard", help="Launch the grade dashboard")

    args = parser.parse_args()

    if args.command == "init":
        # Programmatic validation for init arguments
        if not args.input and (not args.course_name or not args.term or not args.course_id):
            parser.error("Either --input / -i, or all of (--course_name, --term, --course_id) must be specified.")
        init_course(args.course_name, args.term, args.course_id, args.input)
    elif args.command == "mkdb":
        mkdb_course(args.config_file, args.input)
    elif args.command == "update":
        if args.input:
            update_course(input_file=args.input)
        else:
            update_course(config_file=args.config_file)
    elif args.command == "undo":
        undo_course()
    elif args.command == "dashboard":
        run_dashboard()
    else:
        # Default behavior: show help
        parser.print_help()
        sys.exit(1)
