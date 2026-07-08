import sys
from pathlib import Path

# Ensure project root is in the Python search path to support global calls
script_path = Path(__file__).resolve()
project_root = script_path.parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

import json
import math
import re
import yaml
import pandas as pd
from datetime import datetime, timedelta

from src.data_loader import load_config, load_course_data, parse_pts, parse_config_col
from src.calculators import calculate_final_grades, validate_scores, assign_letter_grade
from src.dashboard import export_reports, update_database_totals

def find_courses_fallback() -> dict:
    current_dir = Path(".")
    course_map = {}
    
    # 0. Check if the current directory itself is a course folder
    info_dir = current_dir / "course_info"
    if info_dir.is_dir():
        config_files = [f for f in info_dir.iterdir() if f.is_file() and f.name.endswith("_config.yaml")]
        if config_files:
            course_map[current_dir.resolve().name] = current_dir

    # 1. Scan subdirectories in CWD
    courses_dir = current_dir / "courses"
    potential = []
    if current_dir.is_dir():
        potential.extend([d for d in current_dir.iterdir() if d.is_dir()])
    if courses_dir.is_dir():
        potential.extend([d for d in courses_dir.iterdir() if d.is_dir()])
        
    for d in potential:
        if d.name == "course_info" or d.name == "courses":
            continue
        d_info_dir = d / "course_info"
        if d_info_dir.is_dir():
            config_files = [f for f in d_info_dir.iterdir() if f.is_file() and f.name.endswith("_config.yaml")]
            if config_files:
                course_map[d.name] = d
                
    # 2. Fall back to scanning the project root directory
    if not course_map:
        # Check if project root itself is a course
        proj_info_dir = project_root / "course_info"
        if proj_info_dir.is_dir():
            config_files = [f for f in proj_info_dir.iterdir() if f.is_file() and f.name.endswith("_config.yaml")]
            if config_files:
                course_map[project_root.resolve().name] = project_root

        potential = [d for d in project_root.iterdir() if d.is_dir()]
        proj_courses_dir = project_root / "courses"
        if proj_courses_dir.is_dir():
            potential.extend([d for d in proj_courses_dir.iterdir() if d.is_dir()])
            
        for d in potential:
            if d.name == "course_info" or d.name == "courses":
                continue
            d_info_dir = d / "course_info"
            if d_info_dir.is_dir():
                config_files = [f for f in d_info_dir.iterdir() if f.is_file() and f.name.endswith("_config.yaml")]
                if config_files:
                    course_map[d.name] = d
                    
    return course_map

def get_courses():
    try:
        courses = find_courses_fallback()
        return {
            "status": "success",
            "courses": [{"name": name, "path": str(path.resolve())} for name, path in sorted(courses.items())]
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}

_THAI_DAY = {'จ.': 0, 'อ.': 1, 'พ.': 2, 'พฤ.': 3, 'ศ.': 4, 'ส.': 5, 'อา.': 6}

def _compute_attendance_labels(config: dict, attendance_cols: list) -> dict:
    short_cols = sorted(
        [c for c in attendance_cols if re.match(r'^a\d+$', c)],
        key=lambda c: int(c[1:])
    )
    if not short_cols:
        return {}
    term_start = config.get('term_start_date', '') or ''
    term_start = term_start.strip()
    if not term_start:
        return {}
    day_str = config.get('class_schedule', {}).get('day', '')
    target_wd = _THAI_DAY.get(day_str)
    if target_wd is None:
        return {}
    try:
        start = datetime.strptime(term_start, '%Y-%m-%d')
    except ValueError:
        return {}
    days_ahead = (target_wd - start.weekday()) % 7
    first_class = start + timedelta(days=days_ahead)

    n = len(short_cols)
    weekly = [first_class + timedelta(weeks=i) for i in range(n)]

    extra = []
    for d_str in config.get('attendance_extra_dates', []):
        try:
            extra.append(datetime.strptime(d_str, '%Y-%m-%d'))
        except ValueError:
            pass

    pool = sorted(set(weekly + extra))[:n]

    labels = {}
    for col, d in zip(short_cols, pool):
        labels[col] = f"{d.day} {d.strftime('%b')} {d.year}"
    return labels

def _fill_default_max_scores(config: dict, max_scores: dict) -> None:
    """
    For any data_mapping column without an explicit max score, derive it from the
    category weight: target_total = weight * 100, distributed equally among unspecified
    columns after subtracting the sum of already-specified ones.
    """
    weights = config.get('weights', {})
    data_mapping = config.get('data_mapping', {})
    for category, cols in data_mapping.items():
        if category == 'attendance':
            continue
        weight = float(weights.get(category, 0.0))
        if weight <= 0.0 or not cols:
            continue
        target_total = weight * 100.0
        specified_sum = sum(max_scores[c] for c in cols if c in max_scores)
        unspecified = [c for c in cols if c not in max_scores]
        if unspecified:
            remaining = max(target_total - specified_sum, 0.0)
            per_col = remaining / len(unspecified)
            for col in unspecified:
                max_scores[col] = per_col


_ATTENDANCE_MAP = {'P': 1.0, 'L': 0.8, 'EA': 1.0, 'X': 1.0, 'A': 0.0}


def _compute_analytics(final_df: pd.DataFrame, config: dict, max_scores: dict) -> dict:
    """
    Computes the payload for the Analytics tab. Row order (and therefore each
    student's positional `raw_index`) matches the `student_grades`/`raw_scores`
    arrays built later in get_course_data, since both are derived from the same
    already-sorted final_df in iteration order.
    """
    df = final_df.reset_index(drop=True)
    weights = config.get("weights", {})
    data_mapping = config.get("data_mapping", {})
    n = len(df)

    final_scores = pd.to_numeric(df.get("Final Score", pd.Series(dtype=float)), errors="coerce").fillna(0.0)
    grades = df.get("Grade", pd.Series(dtype=str)).astype(str)

    # ---- Overview ----
    passing_count = int((grades != "F").sum()) if n else 0
    overview = {
        "mean": round(float(final_scores.mean()), 2) if n else 0.0,
        "median": round(float(final_scores.median()), 2) if n else 0.0,
        "std_dev": round(float(final_scores.std(ddof=1)), 2) if n > 1 else 0.0,
        "pass_rate": round(float(passing_count / n * 100), 2) if n else 0.0,
        "passing_count": passing_count,
        "total_count": n,
    }

    # ---- Histogram (Final Score, fixed 0-100 scale, 10 bins) ----
    histogram = []
    if n:
        histogram = [
            {"label": f"{i * 10}-{i * 10 + 9}" if i < 9 else "90-100", "count": 0}
            for i in range(10)
        ]
        for score in final_scores:
            idx = min(9, max(0, int(score // 10)))
            histogram[idx]["count"] += 1

    # ---- Box plot (Final Score, 1.5*IQR outlier rule) ----
    if n:
        q1 = float(final_scores.quantile(0.25))
        q2 = float(final_scores.quantile(0.5))
        q3 = float(final_scores.quantile(0.75))
        iqr = q3 - q1
        lower_fence, upper_fence = q1 - 1.5 * iqr, q3 + 1.5 * iqr
        mean = float(final_scores.mean())

        non_outliers = final_scores[(final_scores >= lower_fence) & (final_scores <= upper_fence)]
        whisker_low = float(non_outliers.min()) if not non_outliers.empty else float(final_scores.min())
        whisker_high = float(non_outliers.max()) if not non_outliers.empty else float(final_scores.max())

        diff_ratio = ((mean - q2) / iqr) if iqr > 0 else 0.0
        if abs(diff_ratio) < 0.05:
            skew_label = "roughly symmetric"
        elif diff_ratio > 0:
            skew_label = "right-skewed (mean > median)"
        else:
            skew_label = "left-skewed (mean < median)"

        outliers = []
        for i in range(n):
            score = float(final_scores.iloc[i])
            if score < lower_fence or score > upper_fence:
                outliers.append({
                    "student_id": str(df.at[i, "Student ID"]),
                    "name": str(df.at[i, "Name"]),
                    "final_score": score,
                    "raw_index": i,
                })
        outliers.sort(key=lambda o: o["final_score"])

        box_plot = {
            "min": float(final_scores.min()), "q1": q1, "median": q2, "q3": q3,
            "max": float(final_scores.max()), "mean": round(mean, 2),
            "whisker_low": whisker_low, "whisker_high": whisker_high,
            "skew_label": skew_label, "outliers": outliers,
        }
    else:
        box_plot = {
            "min": 0.0, "q1": 0.0, "median": 0.0, "q3": 0.0, "max": 0.0, "mean": 0.0,
            "whisker_low": 0.0, "whisker_high": 0.0, "skew_label": "roughly symmetric",
            "outliers": [],
        }

    # ---- Progress over time (class-average per item, in item order) ----
    progress = []
    for category in weights.keys():
        cols = [c for c in data_mapping.get(category, []) if c in df.columns]
        if not cols:
            continue
        items = []
        if category == "attendance":
            for col in cols:
                vals = df[col].astype(str).str.strip().map(_ATTENDANCE_MAP)
                avg_pct = float(vals.mean(skipna=True) * 100) if vals.notna().any() else 0.0
                items.append({"label": col, "avg_pct": round(avg_pct, 2)})
        else:
            for col in cols:
                col_max = max_scores.get(col, 100.0)
                vals = pd.to_numeric(df[col], errors="coerce")
                avg_pct = float(vals.mean(skipna=True) / col_max * 100) if col_max > 0 and vals.notna().any() else 0.0
                items.append({"label": col, "avg_pct": round(avg_pct, 2)})
        progress.append({"category": category, "items": items})

    # ---- Item difficulty & discrimination (categories with >=2 non-attendance items) ----
    item_analysis = []
    for category in weights.keys():
        if category == "attendance":
            continue
        cols = [c for c in data_mapping.get(category, []) if c in df.columns]
        if len(cols) < 2:
            continue
        cat_pct = pd.to_numeric(df.get(f"{category}_pct", pd.Series(dtype=float)), errors="coerce").fillna(0.0)
        for col in cols:
            col_max = max_scores.get(col, 100.0)
            vals = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
            difficulty = float(vals.mean() / col_max * 100) if col_max > 0 else 0.0
            item_share = (vals / col_max * 100) if col_max > 0 else vals * 0.0
            corrected_total = cat_pct - item_share
            discrimination = vals.corr(corrected_total)
            discrimination = float(discrimination) if pd.notna(discrimination) else 0.0
            if difficulty >= 80:
                label = "Easy"
            elif difficulty >= 30:
                label = "Medium"
            else:
                label = "Hard"
            item_analysis.append({
                "category": category, "item": col, "max_score": float(col_max),
                "difficulty": round(difficulty, 2), "difficulty_label": label,
                "discrimination": round(discrimination, 3), "n": int(len(vals)),
            })

    # ---- Correlation matrix (category _pct columns) ----
    pct_cols = [f"{cat}_pct" for cat in weights.keys() if f"{cat}_pct" in df.columns]
    correlation = {"categories": [], "matrix": []}
    if pct_cols:
        corr_df = df[pct_cols].apply(pd.to_numeric, errors="coerce").corr().fillna(0.0)
        correlation["categories"] = [c[:-len("_pct")] for c in pct_cols]
        correlation["matrix"] = [[round(float(v), 3) for v in row] for row in corr_df.values]

    # ---- At-risk students (below Q1 and/or failing) ----
    at_risk = []
    if n:
        q1_final = float(final_scores.quantile(0.25))
        for i in range(n):
            score = float(final_scores.iloc[i])
            grade = str(df.at[i, "Grade"])
            reasons = []
            if score < q1_final:
                reasons.append("Below Q1")
            if grade == "F":
                reasons.append("Failing")
            if reasons:
                at_risk.append({
                    "raw_index": i,
                    "student_id": str(df.at[i, "Student ID"]),
                    "name": str(df.at[i, "Name"]),
                    "final_score": score,
                    "grade": grade,
                    "reasons": reasons,
                })
        at_risk.sort(key=lambda r: r["final_score"])

    return {
        "overview": overview,
        "histogram": histogram,
        "box_plot": box_plot,
        "progress": progress,
        "item_analysis": item_analysis,
        "correlation": correlation,
        "at_risk": at_risk,
    }


def get_course_data(course_path, use_weighted=True):
    try:
        path = Path(course_path)
        config = load_config(path)
        raw_df, max_scores = load_course_data(path)

        if raw_df.empty:
            return {
                "status": "error",
                "message": f"No student data found in {path / 'data'}"
            }

        _fill_default_max_scores(config, max_scores)
        warnings = validate_scores(raw_df, config, max_scores)
        final_df = calculate_final_grades(raw_df, config, max_scores, use_weighted)
        
        # Write totals back to the raw CSV databases
        update_database_totals(path, final_df, config.get("data_mapping", {}), max_scores)
        
        # Prepare list of student grades (summary table)
        data_mapping = config.get("data_mapping", {})
        raw_assignment_cols = {col for cols in data_mapping.values() for col in cols}
        
        weights = config.get("weights", {})
        summary_cols = ["Student ID", "Name"]
        for cat in weights.keys():
            col_name = f"{cat}_pct"
            if col_name in final_df.columns:
                summary_cols.append(col_name)
        if "Final Score" in final_df.columns:
            summary_cols.append("Final Score")
        if "Grade" in final_df.columns:
            summary_cols.append("Grade")
        
        student_grades = []
        for _, row in final_df[summary_cols].iterrows():
            record = {}
            for col in summary_cols:
                val = row[col]
                record[col] = "" if pd.isna(val) else val
            student_grades.append(record)
            
        # Prepare list of raw scores
        raw_cols_present = [c for c in final_df.columns if c in raw_assignment_cols]
        raw_display_cols = ["Student ID", "Name"] + raw_cols_present
        raw_scores = []
        for _, row in final_df[raw_display_cols].iterrows():
            record = {}
            for col in raw_display_cols:
                val = row[col]
                record[col] = "" if pd.isna(val) else val
            raw_scores.append(record)
            
        # Prepare grade distribution
        grade_order = list(config.get("grade_boundaries", {}).keys()) + ["F"]
        counts = final_df["Grade"].value_counts()
        total = len(final_df)
        distribution = {}
        for g in grade_order:
            count = int(counts.get(g, 0))
            pct = float(count / total * 100) if total > 0 else 0.0
            distribution[g] = {"count": count, "pct": pct}
            
        # Prepare roundup summary
        roundup_summary = {
            "improved_count": 0,
            "distribution": [],
            "improved_students": []
        }
        
        if "Original Grade" in final_df.columns:
            orig_counts = final_df["Original Grade"].value_counts().reindex(grade_order, fill_value=0)
            new_counts = final_df["Grade"].value_counts().reindex(grade_order, fill_value=0)
            
            for grade in grade_order:
                orig = int(orig_counts[grade])
                rounded = int(new_counts[grade])
                if orig == 0 and rounded == 0:
                    continue
                roundup_summary["distribution"].append({
                    "grade": grade,
                    "original": orig,
                    "rounded": rounded,
                    "change": rounded - orig
                })
                
            improved = final_df[final_df["Grade"] != final_df["Original Grade"]]
            roundup_summary["improved_count"] = len(improved)
            
            for _, row in improved.iterrows():
                roundup_summary["improved_students"].append({
                    "Student ID": str(row.get("Student ID", "")),
                    "Name": str(row.get("Name", "")),
                    "Original Final Score": float(row.get("Original Final Score", 0)),
                    "Final Score": float(row.get("Final Score", 0)),
                    "Original Grade": str(row.get("Original Grade", "")),
                    "Grade": str(row.get("Grade", ""))
                })
                
        # Clean weights for display/serialization
        clean_weights = {}
        for cat, w in config.get("weights", {}).items():
            clean_weights[cat] = float(w)
            
        # Clean boundaries
        clean_boundaries = {}
        for g, val in config.get("grade_boundaries", {}).items():
            clean_boundaries[g] = float(val)

        # Convert max_scores to floats for serialization
        clean_max_scores = {col: float(val) for col, val in max_scores.items()}

        # Compute human-readable date labels for short attendance columns (a1, a2, …)
        attendance_cols = config.get('data_mapping', {}).get('attendance', []) or []
        attendance_labels = _compute_attendance_labels(config, attendance_cols)

        analytics = _compute_analytics(final_df, config, max_scores)

        return {
            "status": "success",
            "course_id": str(config.get("course_id", "")),
            "course_name": str(config.get("course_name", "")),
            "term": str(config.get("term", "")),
            "weights": clean_weights,
            "grade_boundaries": clean_boundaries,
            "data_mapping": config.get("data_mapping", {}),
            "warnings": warnings,
            "max_scores": clean_max_scores,
            "summary_columns": summary_cols,
            "student_grades": student_grades,
            "raw_columns": raw_display_cols,
            "raw_scores": raw_scores,
            "grade_distribution": distribution,
            "roundup_summary": roundup_summary,
            "rules": config.get("rules", {}),
            "attendance_labels": attendance_labels,
            "analytics": analytics,
        }
    except Exception as e:
        import traceback
        return {
            "status": "error",
            "message": str(e),
            "traceback": traceback.format_exc()
        }

def _update_xlsx_attendance(xlsx_path, student_id, target_col_orig_name, val_to_set):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    
    header_row = 1
    col_idx = None
    student_id_col_idx = None
    
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=c).value
        if val:
            val_str = str(val).strip()
            if val_str == target_col_orig_name:
                col_idx = c
            if val_str == "Student ID":
                student_id_col_idx = c
                
    if not col_idx or not student_id_col_idx:
        return False
        
    row_idx = None
    for r in range(2, ws.max_row + 1):
        sid_val = ws.cell(row=r, column=student_id_col_idx).value
        if sid_val and str(sid_val).strip() == student_id:
            row_idx = r
            break
            
    if not row_idx:
        return False
        
    ws.cell(row=row_idx, column=col_idx).value = val_to_set
    wb.save(xlsx_path)
    return True

def update_student_score(course_path, student_id, col_name, value):
    try:
        path = Path(course_path)
        data_dir = path / "data"
        if not data_dir.is_dir():
            return {"status": "error", "message": f"Data directory not found in {course_path}"}
            
        student_id = str(student_id).strip()
        
        # Try to convert value to appropriate type (float or int if numeric)
        trimmed_val = str(value).strip()
        val_to_set = trimmed_val
        if trimmed_val == "":
            val_to_set = None
        else:
            try:
                if "." in trimmed_val:
                    val_to_set = float(trimmed_val)
                else:
                    val_to_set = int(trimmed_val)
            except ValueError:
                # Keep as string (e.g. attendance codes P, A, L, EA)
                val_to_set = trimmed_val
                
        # Find which file contains the column
        target_file = None
        is_xlsx = False
        target_col_orig_name = None
        
        for file in data_dir.iterdir():
            if file.is_file() and file.suffix in ['.csv', '.xlsx']:
                if file.suffix == '.csv':
                    try:
                        df = pd.read_csv(file, nrows=0)
                        df.columns = df.columns.astype(str).str.strip()
                        for col in df.columns:
                            clean_col, _ = parse_pts(col)
                            if clean_col == col_name or col == col_name:
                                target_file = file
                                target_col_orig_name = col
                                is_xlsx = False
                                break
                    except Exception:
                        pass
                else:
                    try:
                        df = pd.read_excel(file, nrows=0)
                        df.columns = df.columns.astype(str).str.strip()
                        for col in df.columns:
                            clean_col, _ = parse_pts(col)
                            if clean_col == col_name or col == col_name:
                                target_file = file
                                target_col_orig_name = col
                                is_xlsx = True
                                break
                    except Exception:
                        pass
            if target_file:
                break
                
        if not target_file:
            return {"status": "error", "message": f"Column '{col_name}' not found in any score sheets."}
            
        if not is_xlsx:
            # Update CSV file
            df = pd.read_csv(target_file)
            df.columns = df.columns.astype(str).str.strip()
            df['Student ID'] = df['Student ID'].astype(str).str.strip()
            
            mask = df['Student ID'] == student_id
            if not mask.any():
                return {"status": "error", "message": f"Student ID '{student_id}' not found in {target_file.name}."}
                
            if target_col_orig_name in df.columns:
                df[target_col_orig_name] = df[target_col_orig_name].astype(object)
            df.loc[mask, target_col_orig_name] = val_to_set
            
            # Save back
            df.to_csv(target_file, index=False)
            
            # If this is attendance, also update the XLSX companion
            msg = f"Updated {col_name} to {value} for student {student_id} in {target_file.name}."
            if target_file.name.endswith("attendance.csv"):
                xlsx_path = target_file.with_suffix(".xlsx")
                if xlsx_path.exists():
                    ok = _update_xlsx_attendance(xlsx_path, student_id, target_col_orig_name, val_to_set)
                    if ok:
                        msg += " (also updated Excel spreadsheet)"
                    else:
                        msg += " (failed to update Excel spreadsheet)"
            return {"status": "success", "message": msg}
        else:
            # Update XLSX file using openpyxl to preserve formulas/formats
            import openpyxl
            wb = openpyxl.load_workbook(target_file)
            ws = wb.active
            
            # Find column index for target_col_orig_name
            header_row = 1
            col_idx = None
            student_id_col_idx = None
            
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=header_row, column=c).value
                if val:
                    val_str = str(val).strip()
                    if val_str == target_col_orig_name:
                        col_idx = c
                    if val_str == "Student ID":
                        student_id_col_idx = c
                        
            if not col_idx:
                return {"status": "error", "message": f"Column '{target_col_orig_name}' header not found in XLSX sheet."}
            if not student_id_col_idx:
                return {"status": "error", "message": "'Student ID' column header not found in XLSX sheet."}
                
            # Find student row
            row_idx = None
            for r in range(2, ws.max_row + 1):
                sid_val = ws.cell(row=r, column=student_id_col_idx).value
                if sid_val and str(sid_val).strip() == student_id:
                    row_idx = r
                    break
                    
            if not row_idx:
                return {"status": "error", "message": f"Student ID '{student_id}' not found in XLSX sheet."}
                
            # Write cell and save
            ws.cell(row=row_idx, column=col_idx).value = val_to_set
            wb.save(target_file)
            return {"status": "success", "message": f"Updated {col_name} to {value} for student {student_id} in XLSX {target_file.name}."}
            
    except Exception as e:
        import traceback
        return {"status": "error", "message": str(e), "traceback": traceback.format_exc()}

def _update_xlsx_attendance_bulk(xlsx_path, target_col_orig_name, val_to_set):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    header_row = 1
    col_idx = None
    student_id_col_idx = None

    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=c).value
        if val:
            val_str = str(val).strip()
            if val_str == target_col_orig_name:
                col_idx = c
            if val_str == "Student ID":
                student_id_col_idx = c

    if not col_idx or not student_id_col_idx:
        return False

    updated = 0
    for r in range(2, ws.max_row + 1):
        sid_val = ws.cell(row=r, column=student_id_col_idx).value
        if sid_val and str(sid_val).strip():
            ws.cell(row=r, column=col_idx).value = val_to_set
            updated += 1

    if updated == 0:
        return False

    wb.save(xlsx_path)
    return True

def update_column_score(course_path, col_name, value):
    try:
        path = Path(course_path)
        data_dir = path / "data"
        if not data_dir.is_dir():
            return {"status": "error", "message": f"Data directory not found in {course_path}"}

        # Try to convert value to appropriate type (float or int if numeric)
        trimmed_val = str(value).strip()
        val_to_set = trimmed_val
        if trimmed_val == "":
            val_to_set = None
        else:
            try:
                if "." in trimmed_val:
                    val_to_set = float(trimmed_val)
                else:
                    val_to_set = int(trimmed_val)
            except ValueError:
                # Keep as string (e.g. attendance codes P, A, L, EA)
                val_to_set = trimmed_val

        # Find which file contains the column
        target_file = None
        is_xlsx = False
        target_col_orig_name = None

        for file in data_dir.iterdir():
            if file.is_file() and file.suffix in ['.csv', '.xlsx']:
                if file.suffix == '.csv':
                    try:
                        df = pd.read_csv(file, nrows=0)
                        df.columns = df.columns.astype(str).str.strip()
                        for col in df.columns:
                            clean_col, _ = parse_pts(col)
                            if clean_col == col_name or col == col_name:
                                target_file = file
                                target_col_orig_name = col
                                is_xlsx = False
                                break
                    except Exception:
                        pass
                else:
                    try:
                        df = pd.read_excel(file, nrows=0)
                        df.columns = df.columns.astype(str).str.strip()
                        for col in df.columns:
                            clean_col, _ = parse_pts(col)
                            if clean_col == col_name or col == col_name:
                                target_file = file
                                target_col_orig_name = col
                                is_xlsx = True
                                break
                    except Exception:
                        pass
            if target_file:
                break

        if not target_file:
            return {"status": "error", "message": f"Column '{col_name}' not found in any score sheets."}

        if not is_xlsx:
            # Update CSV file
            df = pd.read_csv(target_file)
            df.columns = df.columns.astype(str).str.strip()
            df['Student ID'] = df['Student ID'].astype(str).str.strip()

            mask = df['Student ID'].str.len() > 0
            student_count = int(mask.sum())

            if target_col_orig_name in df.columns:
                df[target_col_orig_name] = df[target_col_orig_name].astype(object)
            df.loc[mask, target_col_orig_name] = val_to_set

            # Save back
            df.to_csv(target_file, index=False)

            msg = f"Filled '{col_name}' = {value} for {student_count} students in {target_file.name}."
            if target_file.name.endswith("attendance.csv"):
                xlsx_path = target_file.with_suffix(".xlsx")
                if xlsx_path.exists():
                    ok = _update_xlsx_attendance_bulk(xlsx_path, target_col_orig_name, val_to_set)
                    if ok:
                        msg += " (also updated Excel spreadsheet)"
                    else:
                        msg += " (failed to update Excel spreadsheet)"
            return {"status": "success", "message": msg}
        else:
            # Update XLSX file using openpyxl to preserve formulas/formats
            import openpyxl
            wb = openpyxl.load_workbook(target_file)
            ws = wb.active

            header_row = 1
            col_idx = None
            student_id_col_idx = None

            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=header_row, column=c).value
                if val:
                    val_str = str(val).strip()
                    if val_str == target_col_orig_name:
                        col_idx = c
                    if val_str == "Student ID":
                        student_id_col_idx = c

            if not col_idx:
                return {"status": "error", "message": f"Column '{target_col_orig_name}' header not found in XLSX sheet."}
            if not student_id_col_idx:
                return {"status": "error", "message": "'Student ID' column header not found in XLSX sheet."}

            updated = 0
            for r in range(2, ws.max_row + 1):
                sid_val = ws.cell(row=r, column=student_id_col_idx).value
                if sid_val and str(sid_val).strip():
                    ws.cell(row=r, column=col_idx).value = val_to_set
                    updated += 1

            wb.save(target_file)
            return {"status": "success", "message": f"Filled '{col_name}' = {value} for {updated} students in XLSX {target_file.name}."}

    except Exception as e:
        import traceback
        return {"status": "error", "message": str(e), "traceback": traceback.format_exc()}

def update_course_config(course_path, weights_dict=None, boundaries_dict=None):
    try:
        path = Path(course_path)
        info_dir = path / "course_info"
        if not info_dir.is_dir():
            return {"status": "error", "message": f"course_info directory not found in {course_path}"}
            
        config_files = [f for f in info_dir.iterdir() if f.is_file() and f.name.endswith("_config.yaml")]
        if not config_files:
            return {"status": "error", "message": f"No config YAML found in {info_dir}"}
            
        config_file = config_files[0]
        
        with open(config_file, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f) or {}
            
        # Update weights
        if weights_dict is not None:
            config["weights"] = weights_dict
            
        # Update grade boundaries
        if boundaries_dict is not None:
            # boundaries should be floats, sort highest first
            cleaned_bounds = {}
            for k, v in boundaries_dict.items():
                try:
                    cleaned_bounds[str(k).strip()] = float(v)
                except ValueError:
                    pass
            # Sort bounds
            sorted_bounds = sorted(cleaned_bounds.items(), key=lambda x: x[1], reverse=True)
            config["grade_boundaries"] = {k: v for k, v in sorted_bounds}
            
        with open(config_file, 'w', encoding='utf-8') as f:
            yaml.dump(config, f, allow_unicode=True, default_flow_style=False, sort_keys=False)
            
        return {"status": "success", "message": "Updated configuration successfully."}
    except Exception as e:
        return {"status": "error", "message": str(e)}

def run_export(course_path, use_weighted=True):
    try:
        path = Path(course_path)
        config = load_config(path)
        raw_df, max_scores = load_course_data(path)
        final_df = calculate_final_grades(raw_df, config, max_scores, use_weighted)
        export_reports(final_df, path, config, max_scores, use_weighted)
        return {"status": "success", "message": f"Reports successfully exported to {path}/reports/"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

if __name__ == "__main__":
    # Command line runner for debugging
    if len(sys.argv) < 2:
        print(json.dumps({"status": "error", "message": "No command provided"}))
        sys.exit(1)
        
    cmd = sys.argv[1]
    
    if cmd == "get-courses":
        print(json.dumps(get_courses()))
    elif cmd == "get-course-data":
        if len(sys.argv) < 3:
            print(json.dumps({"status": "error", "message": "Missing course path"}))
            sys.exit(1)
        course_path = sys.argv[2]
        use_weighted = True
        if len(sys.argv) > 3:
            use_weighted = sys.argv[3].lower() in ['true', '1', 'yes']
        print(json.dumps(get_course_data(course_path, use_weighted)))
    elif cmd == "update-score":
        if len(sys.argv) < 6:
            print(json.dumps({"status": "error", "message": "Missing arguments (path, student_id, col, val)"}))
            sys.exit(1)
        course_path = sys.argv[2]
        student_id = sys.argv[3]
        col_name = sys.argv[4]
        value = sys.argv[5]
        print(json.dumps(update_student_score(course_path, student_id, col_name, value)))
    elif cmd == "bulk-update-score":
        if len(sys.argv) < 5:
            print(json.dumps({"status": "error", "message": "Missing arguments (path, col, val)"}))
            sys.exit(1)
        course_path = sys.argv[2]
        col_name = sys.argv[3]
        value = sys.argv[4]
        print(json.dumps(update_column_score(course_path, col_name, value)))
    elif cmd == "update-config":
        if len(sys.argv) < 5:
            print(json.dumps({"status": "error", "message": "Missing arguments (path, weights_json, boundaries_json)"}))
            sys.exit(1)
        course_path = sys.argv[2]
        weights_dict = json.loads(sys.argv[3]) if sys.argv[3] and sys.argv[3] != "null" else None
        boundaries_dict = json.loads(sys.argv[4]) if sys.argv[4] and sys.argv[4] != "null" else None
        print(json.dumps(update_course_config(course_path, weights_dict, boundaries_dict)))
    elif cmd == "export-reports":
        if len(sys.argv) < 3:
            print(json.dumps({"status": "error", "message": "Missing course path"}))
            sys.exit(1)
        course_path = sys.argv[2]
        use_weighted = True
        if len(sys.argv) > 3:
            use_weighted = sys.argv[3].lower() in ['true', '1', 'yes']
        print(json.dumps(run_export(course_path, use_weighted)))
    else:
        print(json.dumps({"status": "error", "message": f"Unknown command {cmd}"}))
